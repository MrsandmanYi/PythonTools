
# python-docx
import sys
from docx import Document
from openpyxl import Workbook
#from translate import Translator
import os
import re
import glob

def contains_chinese(s):
    return bool(re.search('[\u4e00-\u9fff]', s))

def read_word_file_lines(file_path):
    doc = Document(file_path)
    full_text = []
    for para in doc.paragraphs:
        full_text.append(para.text)
    return full_text



class ExcelElement:
    def __init__(self):
        self.enWord = ''
        self.cnWord = ''

# 使用函数
file_path = ""
if sys.argv and len(sys.argv) > 1 and len(sys.argv[1]) > 0:
    file_path = sys.argv[1]
else:
    file_path = ""

    # Get the current script's directory
    script_directory = os.path.dirname(os.path.abspath(__file__))

    # Get all the word files in the script's directory
    word_files = glob.glob(os.path.join(script_directory, '*.docx'))

    # Check if there are any word files
    if word_files:
        # Use the first word file found
        file_path = word_files[0]

print("file_path: " + file_path)
textLines = read_word_file_lines(file_path)
lineCount = len(textLines)

excelElements = []

currentLine = 0

while currentLine < lineCount:
    line = textLines[currentLine]
    line = line.lower()
    line = line.strip()
    if len(line) == 0 or contains_chinese(line) or (line.count('/') > 0):
        currentLine += 1
        continue
    
    excelElement = ExcelElement()
    excelElement.enWord = line
    
    findCnWord = False
    while not findCnWord:
        currentLine += 1
        line = textLines[currentLine]
        line = line.lower()
        line = line.strip()
        
        if len(line) == 0 or (line.count('/') > 0):
            continue
        else:
            excelElement.cnWord = line
            findCnWord = True
    
    excelElements.append(excelElement)
    currentLine += 1

wb = Workbook()
ws = wb.active
ws['A1'] = 'enWord'
ws['B1'] = 'cnWord'
#ws['C1'] = 'cnWord2'

txtFileLines = []

# 创建一个翻译器，设置源语言和目标语言
#translator = Translator(to_lang = "zh", from_lang = "en")
#print(translator.translate("This is a pen."))
for i, excelElement in enumerate(excelElements):
    ws.cell(row=i+2, column=1, value=excelElement.enWord)
    ws.cell(row=i+2, column=2, value=excelElement.cnWord)
    txtFileLines.append(excelElement.enWord + "\t" + excelElement.cnWord)
    # 通过翻译器翻译
    #cnWord2 = translator.translate(excelElement.enWord)
    #print(i +"/"+len(excelElements) + " "+ excelElement.enWord + " " + cnWord2)
    #ws.cell(row=i+2, column=3, value=cnWord2)

xlsxFileName = file_path.replace('.docx', '.xlsx')
txtFileName = file_path.replace('.docx', '.txt')


wb.save(xlsxFileName)
with open(txtFileName, 'w', encoding='utf-8') as f:
    f.write('\n'.join(txtFileLines))

#os.system('start ' + xlsxFileName)
os.system('start ' + txtFileName)